import os
import json
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime


def find_json_files(folder_path):
    """Найти все JSON файлы в папке"""
    json_files = []
    for file in os.listdir(folder_path):
        if file.lower().endswith('.json'):
            json_files.append(os.path.join(folder_path, file))
    return json_files


def check_duplicates_in_file(file_path):
    """Проверить дубликаты в одном файле"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Получаем характеристики
        characteristics = data.get('ПараметрыТовара', {}).get('Характеристики', [])
        if not characteristics:
            return None

        # Собираем названия характеристик
        names = []
        for char in characteristics:
            name = char.get('Наименование')
            if name:
                names.append(name)

        # Проверяем дубликаты
        counter = Counter(names)
        duplicates = {name: count for name, count in counter.items() if count > 1}

        if duplicates:
            return {
                'file_name': os.path.basename(file_path),
                'file_path': file_path,
                'product_name': data.get('ПараметрыТовара', {}).get('Наименование', ''),
                'total_chars': len(characteristics),
                'duplicates': duplicates,
                'all_names': names
            }
    except Exception as e:
        print(f"Ошибка в файле {os.path.basename(file_path)}: {e}")
    return None


def create_excel_report(duplicates_list, output_file):
    """Создать Excel отчет"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Дубликаты характеристик"

    # Заголовки
    headers = [
        "Имя файла",
        "Название товара",
        "Дублирующаяся характеристика",
        "Количество повторов",
        "Всего характеристик в файле",
        "Путь к файлу"
    ]

    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Заполняем данные
    row = 2
    for file_info in duplicates_list:
        for dup_name, count in file_info['duplicates'].items():
            ws.cell(row=row, column=1, value=file_info['file_name'])
            ws.cell(row=row, column=2, value=file_info['product_name'])
            ws.cell(row=row, column=3, value=dup_name)
            ws.cell(row=row, column=4, value=count)
            ws.cell(row=row, column=5, value=file_info['total_chars'])
            ws.cell(row=row, column=6, value=file_info['file_path'])
            row += 1

    # Настраиваем ширину столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем файл
    wb.save(output_file)
    print(f"✓ Отчет сохранен: {output_file}")


def main():
    """Основная функция"""
    # Путь к папке с JSON файлами
    folder_path = "C:/Users/reus/Desktop/СТИЛбЕрг/УВ/УВ json"

    # Проверяем папку
    if not os.path.exists(folder_path):
        print(f"❌ Папка не существует: {folder_path}")
        return

    print(f"🔍 Сканирую папку: {folder_path}")

    # Ищем все JSON файлы
    json_files = find_json_files(folder_path)
    print(f"📁 Найдено JSON файлов: {len(json_files)}")

    if not json_files:
        print("❌ JSON файлы не найдены")
        return

    # Проверяем каждый файл на дубликаты
    duplicates_list = []
    for i, file_path in enumerate(json_files, 1):
        print(f"   Обработка {i}/{len(json_files)}: {os.path.basename(file_path)}")
        result = check_duplicates_in_file(file_path)
        if result:
            duplicates_list.append(result)

    print(f"\n📊 Результаты:")
    print(f"   Файлов с дубликатами: {len(duplicates_list)}")

    if duplicates_list:
        # Создаем отчет
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = f"duplicates_report_{timestamp}.xlsx"
        create_excel_report(duplicates_list, report_file)

        # Выводим сводку
        print("\n📋 Файлы с дубликатами:")
        for file_info in duplicates_list:
            print(f"   • {file_info['file_name']}")
            print(f"     Товар: {file_info['product_name']}")
            print(f"     Дубликаты: {', '.join([f'{k} ({v} раз)' for k, v in file_info['duplicates'].items()])}")
            print()
    else:
        print("✓ Дубликатов не найдено")


if __name__ == "__main__":
    # Проверяем наличие openpyxl
    try:
        from openpyxl import Workbook
    except ImportError:
        print("❌ Установите библиотеку openpyxl:")
        print("   pip install openpyxl")
        exit(1)

    main()